"""会員名簿 正規化インポートのユニットテスト（合成フィクスチャ）。"""
import openpyxl

from app.models import MemberType
from app.roster_import import apply_org_mapping, parse_members_sheet


def build_fixture_sheet():
    """実レイアウト（1会員=2行・2行ヘッダー）を模した合成シート。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 198  # 上部の余り数値（無視される）
    # 2行ヘッダー
    ws["A2"], ws["B2"], ws["C2"], ws["D2"], ws["E2"], ws["F2"], ws["G2"] = (
        "Ｎｏ", "氏名", "〒", "TEL", "勤務先", "TEL", "Mobile",
    )
    ws["B3"], ws["C3"], ws["E3"], ws["F3"], ws["G3"] = (
        "生年月日", "自宅住所", "役職名", "FAX", "E-Mail",
    )
    # 会員1（全角を含めてNFKC正規化を検証）
    ws["A4"], ws["B4"], ws["C4"], ws["D4"], ws["E4"], ws["F4"], ws["G4"] = (
        1, "遠藤孝行", "９６９‐３１００", "０８０‐３８０１‐５３４９",
        "株式会社 アウレ", "024-548-6051", "080-3801-5349",
    )
    ws["B5"], ws["C5"], ws["E5"], ws["F5"], ws["G5"] = (
        "H3.11.15", "猪苗代町新町4931-1", "代表取締役", "024-548-6051", "taro@example.com",
    )
    # 会員2
    ws["A6"], ws["B6"], ws["C6"], ws["E6"], ws["F6"], ws["G6"] = (
        2, "遠藤大介", "969-3133", "株式会社 鶴我", "0242-23-8294", "080-6041-3846",
    )
    ws["B7"], ws["C7"], ws["E7"], ws["G7"] = (
        "H4.3.6", "猪苗代町大字千代田字柳田11", "社員", "jiro@example.com",
    )
    return ws


def test_parse_basic_fields():
    members = parse_members_sheet(build_fixture_sheet())
    assert len(members) == 2

    m1 = members[0]
    assert m1.member_id == "m1"
    assert m1.name == "遠藤孝行"
    assert m1.birthday == "H3.11.15"
    assert m1.contact.work == "株式会社 アウレ"
    assert m1.contact.work_title == "代表取締役"
    assert m1.contact.home_address == "猪苗代町新町4931-1"
    assert m1.contact.email == "taro@example.com"


def test_nfkc_normalization():
    """全角数字・全角ハイフンが半角へ正規化される。"""
    members = parse_members_sheet(build_fixture_sheet())
    m1 = members[0]
    assert m1.contact.postal_code == "969-3100"
    assert m1.contact.mobile == "080-3801-5349"


def test_second_member_partial_fields():
    members = parse_members_sheet(build_fixture_sheet())
    m2 = members[1]
    assert m2.name == "遠藤大介"
    assert m2.contact.home_tel is None  # 上段D未入力
    assert m2.contact.fax is None  # 下段F未入力
    assert m2.contact.email == "jiro@example.com"


def test_apply_org_mapping():
    members = parse_members_sheet(build_fixture_sheet())
    mapping = {
        "遠藤孝行": {
            "committee": "コト創り委員会",
            "committee_role": "委員長",
            "officer_role": None,
            "member_type": "regular",
        },
        "遠藤大介": {"committee": "総務委員会", "committee_role": "委員"},
    }
    updated = apply_org_mapping(members, mapping)
    assert updated == 2
    assert members[0].committee == "コト創り委員会"
    assert members[0].committee_role == "委員長"
    assert members[0].member_type == MemberType.regular
    assert members[1].committee == "総務委員会"
