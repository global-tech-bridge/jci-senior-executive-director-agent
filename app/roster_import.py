"""会員名簿（カード型 xlsx）の正規化インポート。

実レイアウト（docs/drive-analysis.md §10, 1会員=2行）:
  上段:  A=No  B=氏名      C=〒        D=自宅TEL  E=勤務先    F=勤務先TEL  G=Mobile
  下段:        B=生年月日  C=自宅住所  -          E=役職名    F=FAX        G=E-Mail
ヘッダーは2行（氏名行・生年月日行）。データは header+2 から2行刻み。
"""
from __future__ import annotations

import unicodedata

from .models import Contact, Member, MemberType

# NFKC で正規化されないダッシュ類（電話/郵便番号で混在しがち）を ASCII ハイフンへ
_DASHES = "‐‑‒–—―−"
_DASH_TABLE = {ord(c): "-" for c in _DASHES}


def _norm(value) -> str | None:
    """全角→半角(NFKC)・ダッシュ統一・前後空白除去。空は None。"""
    if value is None:
        return None
    s = unicodedata.normalize("NFKC", str(value)).translate(_DASH_TABLE).strip()
    return s or None


def _name_key(name: str) -> str:
    """氏名マッチ用キー（空白除去）。"""
    return unicodedata.normalize("NFKC", name).replace(" ", "").replace("　", "")


def find_header_row(ws, *, max_scan: int = 15) -> int:
    """col B が「氏名」の行を探す。"""
    for r in range(1, min(max_scan, ws.max_row) + 1):
        if _norm(ws.cell(r, 2).value) == "氏名":
            return r
    raise ValueError("ヘッダー行（氏名）が見つかりません")


def parse_members_sheet(ws, *, lom_id: str = "inawashiro") -> list[Member]:
    header_row = find_header_row(ws)
    members: list[Member] = []
    r = header_row + 2  # 2行ヘッダーの次から
    while r <= ws.max_row:
        name = _norm(ws.cell(r, 2).value)
        if name:
            no = ws.cell(r, 1).value
            no_str = str(no).strip() if no is not None else ""
            if no_str.replace(".0", "").isdigit():
                member_id = f"m{int(float(no_str))}"
            else:
                member_id = f"m{len(members) + 1}"
            contact = Contact(
                postal_code=_norm(ws.cell(r, 3).value),
                home_tel=_norm(ws.cell(r, 4).value),
                home_address=_norm(ws.cell(r + 1, 3).value),
                work=_norm(ws.cell(r, 5).value),
                work_title=_norm(ws.cell(r + 1, 5).value),
                work_tel=_norm(ws.cell(r, 6).value),
                fax=_norm(ws.cell(r + 1, 6).value),
                mobile=_norm(ws.cell(r, 7).value),
                email=_norm(ws.cell(r + 1, 7).value),
            )
            members.append(
                Member(
                    member_id=member_id,
                    lom_id=lom_id,
                    name=name,
                    birthday=_norm(ws.cell(r + 1, 2).value),
                    contact=contact,
                )
            )
        r += 2
    return members


def load_roster(path: str, *, sheet: str | None = None, lom_id: str = "inawashiro") -> list[Member]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet is None:
        # 「会員名簿」を含むシートを優先、なければ先頭
        sheet = next((s for s in wb.sheetnames if "会員名簿" in s), wb.sheetnames[0])
    return parse_members_sheet(wb[sheet], lom_id=lom_id)


def apply_org_mapping(members: list[Member], mapping: dict[str, dict]) -> int:
    """組織図由来の委員会/ロール/会員区分を氏名キーで付与。更新件数を返す。

    mapping 例: {"遠藤孝行": {"committee": "コト創り委員会",
                            "committee_role": "委員長",
                            "officer_role": None,
                            "member_type": "regular"}}
    """
    updated = 0
    for m in members:
        info = mapping.get(_name_key(m.name))
        if not info:
            continue
        if "committee" in info:
            m.committee = info["committee"]
        if "committee_role" in info:
            m.committee_role = info["committee_role"]
        if "officer_role" in info:
            m.officer_role = info["officer_role"]
        if "member_type" in info:
            m.member_type = MemberType(info["member_type"])
        updated += 1
    return updated
